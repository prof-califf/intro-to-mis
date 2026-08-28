#!/usr/bin/env python3
"""Add an Analyst tab to the student workbooks.

Same layout in every lab: one row per checkpoint, answer in column B.
Students never have to decide where to put anything, and the grader
opens the same cell every time.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "0B1D33"
BLUE = "1A5276"
PALE = "F2F6F9"

SOURCES = [
    ("/mnt/project/cascadiasales2025.xlsx", "cascadia-sales-2025.xlsx", "Sales2025"),
    ("/mnt/project/cascadiasales2025RAW.xlsx", "cascadia-sales-2025-RAW.xlsx", "Sales2025_RAW"),
]

ROWS = 10  # Lab 4 and Lab 5 have 8 checkpoints; 10 covers every lab

def build(src, out, datasheet):
    wb = openpyxl.load_workbook(src)
    if "Analyst" in wb.sheetnames:
        del wb["Analyst"]
    ws = wb.create_sheet("Analyst", 0)  # first tab, so it opens here

    ws["A1"] = "Your answers"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)

    ws["A2"] = (f"Put every checkpoint answer in column B, on the matching row. "
                f"The data is on the {datasheet} tab.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="5A5A5A")
    ws.merge_cells("A2:C2")

    hdr = ["Checkpoint", "Your answer or formula", "Notes (optional)"]
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="C9D3DC")
    for r in range(5, 5 + ROWS):
        n = r - 4
        a = ws.cell(row=r, column=1, value=n)
        a.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        a.alignment = Alignment(horizontal="center")
        a.fill = PatternFill("solid", fgColor=PALE)
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).border = Border(
                top=thin, bottom=thin, left=thin, right=thin)
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=11)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 40
    for r in range(5, 5 + ROWS):
        ws.row_dimensions[r].height = 22
    ws.sheet_view.showGridLines = False

    wb.save(out)
    return wb.sheetnames

for src, out, ds in SOURCES:
    names = build(src, out, ds)
    print(f"{out}: sheets = {names}")
